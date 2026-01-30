"""Main application window with three-panel layout.

This module provides the MainWindow class that serves as the main
application window, containing the CardDesigner, CardTableWidget, and
DetailsBar in a splitter layout.
"""

import json
import shutil
import uuid
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Qt, QUrl
from PySide6.QtGui import QCloseEvent, QKeySequence, QAction, QDesktopServices
from PySide6.QtWidgets import (
    QMainWindow,
    QSplitter,
    QWidget,
    QVBoxLayout,
    QMessageBox,
    QInputDialog,
    QFileDialog,
    QDialog,
    QListWidget,
    QDialogButtonBox,
    QLabel,
    QComboBox,
    QFormLayout,
)

from src.business_card_generator.core.card_data_model import CardDataModel
from src.business_card_generator.models.card import CardTemplate, CardRow
from src.business_card_generator.ui.card_designer import CardDesigner
from src.business_card_generator.ui.card_table_view import CardTableWidget
from src.business_card_generator.ui.details_bar import DetailsBar


class MainWindow(QMainWindow):
    """Main application window with three-panel layout.
    
    Layout:
    - CardDesigner (left): Visual card preview with draggable fields
    - CardTableWidget (center): Table of card data with add/remove buttons
    - DetailsBar (right): Property editor for selected field
    """
    
    BASE_DIR = Path.home() / ".business-card-generator"
    
    def __init__(self, model: Optional[CardDataModel] = None):
        super().__init__()
        
        self._model = model if model else CardDataModel()
        self._has_unsaved_changes = False
        self._current_project_name: Optional[str] = None
        
        self.setWindowTitle("Business Card Generator")
        self.setMinimumSize(1200, 700)
        
        self._card_designer: Optional[CardDesigner] = None
        self._card_table_widget: Optional[CardTableWidget] = None
        self._details_bar: Optional[DetailsBar] = None
        self._splitter: Optional[QSplitter] = None
        self._menu_actions: dict[str, QAction] = {}
        
        self._setup_ui()
        self._setup_menu()
        self._connect_signals()
    
    def _setup_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        
        layout = QVBoxLayout(central)
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(0)
        
        self._splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Create panels
        self._card_designer = CardDesigner()
        self._card_designer.set_model(self._model)
        
        self._card_table_widget = CardTableWidget(self._model)
        self._details_bar = DetailsBar()
        
        self._splitter.addWidget(self._card_designer)
        self._splitter.addWidget(self._card_table_widget)
        self._splitter.addWidget(self._details_bar)
        
        self._splitter.setSizes([350, 500, 250])
        self._splitter.setCollapsible(0, False)
        self._splitter.setCollapsible(1, False)
        self._splitter.setCollapsible(2, True)
        
        layout.addWidget(self._splitter)
    
    def _setup_menu(self) -> None:
        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("&File")
        
        # New Project
        new_action = QAction("&New Project", self)
        new_action.setShortcut(QKeySequence("Ctrl+N"))
        new_action.setToolTip("Create a new project (Ctrl+N)")
        new_action.triggered.connect(self._on_new_project)
        file_menu.addAction(new_action)
        self._menu_actions["new_project"] = new_action
        
        # Open Project
        open_action = QAction("&Open Project", self)
        open_action.setShortcut(QKeySequence("Ctrl+O"))
        open_action.setToolTip("Open an existing project (Ctrl+O)")
        open_action.triggered.connect(self._on_open_project)
        file_menu.addAction(open_action)
        self._menu_actions["open_project"] = open_action
        
        # Save Project
        save_action = QAction("&Save Project", self)
        save_action.setShortcut(QKeySequence("Ctrl+S"))
        save_action.setToolTip("Save the current project (Ctrl+S)")
        save_action.triggered.connect(self._on_save_project)
        file_menu.addAction(save_action)
        self._menu_actions["save_project"] = save_action
        
        file_menu.addSeparator()
        
        # Import from Excel
        import_action = QAction("&Import from Excel...", self)
        import_action.setShortcut(QKeySequence("Ctrl+I"))
        import_action.setToolTip("Import card data from Excel file (Ctrl+I)")
        import_action.triggered.connect(self._on_import_excel)
        file_menu.addAction(import_action)
        self._menu_actions["import_excel"] = import_action
        
        # Export
        export_action = QAction("&Export...", self)
        export_action.setShortcut(QKeySequence("Ctrl+E"))
        export_action.setToolTip("Export cards to PDF or Word (Ctrl+E)")
        export_action.triggered.connect(self._on_export)
        file_menu.addAction(export_action)
        self._menu_actions["export"] = export_action
        
        file_menu.addSeparator()
        
        # Exit
        exit_action = QAction("E&xit", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.setToolTip("Exit the application (Ctrl+Q)")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        self._menu_actions["exit"] = exit_action
    
    def _connect_signals(self) -> None:
        # Table row selection -> update card preview
        self._card_table_widget.row_selected.connect(self._on_row_selected)
        self._card_table_widget.row_added.connect(self._on_row_added)
        self._card_table_widget.row_removed.connect(self._on_row_removed)
        self._card_table_widget.image_selected.connect(self._on_image_selected)
        
        # Designer field selection -> update details bar
        self._card_designer.field_selected.connect(self._on_field_selected)
        self._card_designer.field_position_changed.connect(self._on_field_position_changed)
        
        # Details bar changes -> update designer
        self._details_bar.field_changed.connect(self._on_field_changed)
        
        # Model changes -> mark unsaved
        self._model.dataChanged.connect(self._mark_unsaved)
        self._model.template_changed.connect(self._mark_unsaved)
    
    def _on_row_selected(self, row_id: str) -> None:
        """Handle row selection - update card preview."""
        row = self._model.get_row(row_id)
        self._card_designer.set_row(row)
    
    def _on_row_added(self, row_id: str) -> None:
        """Handle new row added."""
        self._mark_unsaved()
    
    def _on_row_removed(self, row_id: str) -> None:
        """Handle row removed."""
        self._card_designer.clear()
        self._mark_unsaved()
    
    def _on_image_selected(self, file_path: str, row_index: int, field_id: str) -> None:
        """Handle image selection - copy to project folder if project exists."""
        if not self._current_project_name:
            # No project yet, just use the absolute path for now
            return
        
        # Copy image to project's images folder
        source = Path(file_path)
        if not source.exists():
            return
        
        project_path = self.BASE_DIR / self._current_project_name
        images_dir = project_path / "images"
        images_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename to avoid conflicts
        ext = source.suffix
        unique_name = f"{uuid.uuid4().hex[:8]}_{source.name}"
        dest = images_dir / unique_name
        
        try:
            shutil.copy2(source, dest)
            
            # Update the model with the relative path
            row = self._model.get_row_at_index(row_index)
            if row:
                relative_path = f"images/{unique_name}"
                row.set_value(field_id, relative_path)
                
                # Refresh the designer if this row is selected
                selected_id = self._card_table_widget.get_selected_row_id()
                if selected_id == row.id:
                    self._card_designer.set_row(row)
                
                self._mark_unsaved()
        except OSError as e:
            QMessageBox.warning(self, "Warning", f"Could not copy image: {e}")
    
    def _on_field_selected(self, field_id: str) -> None:
        """Handle field selection in designer - show in details bar."""
        field = self._model.get_field_by_id(field_id)
        if field:
            self._details_bar.set_field(field)
    
    def _on_field_position_changed(self, field_id: str, x: int, y: int) -> None:
        """Handle field position change from designer."""
        # Update details bar if this field is selected
        current = self._details_bar.get_current_field()
        if current and current.id == field_id:
            self._details_bar.set_field(current)
        self._mark_unsaved()
    
    def _on_field_changed(self, field_id: str) -> None:
        """Handle field property change from details bar."""
        # Rebuild designer to reflect changes
        self._card_designer._rebuild_field_widgets()
        row_id = self._card_table_widget.get_selected_row_id()
        if row_id:
            row = self._model.get_row(row_id)
            self._card_designer.set_row(row)
        self._mark_unsaved()
    
    def _mark_unsaved(self) -> None:
        self._has_unsaved_changes = True
        self._update_title()
    
    def _update_title(self) -> None:
        title = "Business Card Generator"
        if self._current_project_name:
            title = f"{self._current_project_name} - Business Card Generator"
        if self._has_unsaved_changes:
            title += " *"
        self.setWindowTitle(title)
    
    def _on_new_project(self) -> None:
        if self._has_unsaved_changes:
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                "You have unsaved changes. Save before creating new project?",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Save
            )
            if reply == QMessageBox.StandardButton.Save:
                if not self._on_save_project():
                    return
            elif reply == QMessageBox.StandardButton.Cancel:
                return
        
        name, ok = QInputDialog.getText(self, "New Project", "Enter project name:")
        if not ok or not name.strip():
            return
        
        name = name.strip()
        project_path = self.BASE_DIR / name
        
        if project_path.exists():
            reply = QMessageBox.question(
                self, "Project Exists",
                f"Project '{name}' already exists. Overwrite?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        # Create new project
        self._model = CardDataModel()
        self._current_project_name = name
        self._has_unsaved_changes = False
        
        # Save immediately to create the project folder
        self._save_to_disk()
        
        # Set project path on designer for image resolution
        self._card_designer.set_project_path(str(self.BASE_DIR / name))
        
        # Update UI
        self._card_designer.set_model(self._model)
        self._card_table_widget._model = self._model
        table_view = self._card_table_widget._table_view
        table_view._model = self._model
        table_view.setModel(self._model)
        
        # Reconnect selection model signals
        sel_model = table_view.selectionModel()
        if sel_model:
            sel_model.selectionChanged.connect(table_view._on_selection_changed)
        
        # Reconnect template changed signal
        self._model.template_changed.connect(table_view._configure_headers)
        table_view._configure_headers()
        
        self._details_bar.clear()
        self._card_designer.clear()
        self._card_table_widget._remove_row_btn.setEnabled(False)
        
        self._model.dataChanged.connect(self._mark_unsaved)
        self._model.template_changed.connect(self._mark_unsaved)
        
        self._update_title()
        self.statusBar().showMessage(f"Project '{name}' created.", 3000)
    
    def _on_open_project(self) -> None:
        if self._has_unsaved_changes:
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                "You have unsaved changes. Save before opening another project?",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Save
            )
            if reply == QMessageBox.StandardButton.Save:
                if not self._on_save_project():
                    return
            elif reply == QMessageBox.StandardButton.Cancel:
                return
        
        # List available projects
        projects = self._list_projects()
        if not projects:
            QMessageBox.information(self, "No Projects", "No existing projects found.")
            return
        
        dialog = OpenProjectDialog(projects, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            selected = dialog.get_selected_project()
            if selected:
                self._load_from_disk(selected)
    
    def _on_save_project(self) -> bool:
        if not self._current_project_name:
            name, ok = QInputDialog.getText(self, "Save Project", "Enter project name:")
            if not ok or not name.strip():
                return False
            self._current_project_name = name.strip()
        
        return self._save_to_disk()
    
    def _save_to_disk(self) -> bool:
        if not self._current_project_name:
            return False
        
        project_path = self.BASE_DIR / self._current_project_name
        project_path.mkdir(parents=True, exist_ok=True)
        
        data = {
            "version": "2.0",
            "template": self._model.get_template().to_dict(),
            "rows": [row.to_dict() for row in self._model.get_all_rows()]
        }
        
        try:
            with open(project_path / "project.json", 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            
            self._has_unsaved_changes = False
            self._update_title()
            self.statusBar().showMessage(f"Project saved.", 3000)
            return True
        except OSError as e:
            QMessageBox.critical(self, "Error", f"Failed to save: {e}")
            return False
    
    def _load_from_disk(self, name: str) -> bool:
        project_path = self.BASE_DIR / name
        project_file = project_path / "project.json"
        
        if not project_file.exists():
            QMessageBox.critical(self, "Error", f"Project file not found.")
            return False
        
        try:
            with open(project_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Load template and rows
            template = CardTemplate.from_dict(data.get("template", {}))
            rows_data = data.get("rows", [])
            
            # Create new model with loaded data
            self._model = CardDataModel()
            self._model._template = template
            for row_data in rows_data:
                row = CardRow.from_dict(row_data)
                self._model._rows.append(row)
            
            self._current_project_name = name
            self._has_unsaved_changes = False
            
            # Set project path on designer for image resolution
            self._card_designer.set_project_path(str(project_path))
            
            # Update designer with new model
            self._card_designer.set_model(self._model)
            
            # Explicitly rebuild field widgets to show the loaded template
            self._card_designer._rebuild_field_widgets()
            
            # Update table widget - need to update both the widget and inner table view
            self._card_table_widget._model = self._model
            table_view = self._card_table_widget._table_view
            table_view._model = self._model
            table_view.setModel(self._model)
            
            # Reconnect selection model signals
            sel_model = table_view.selectionModel()
            if sel_model:
                sel_model.selectionChanged.connect(table_view._on_selection_changed)
            
            # Reconnect template changed signal
            self._model.template_changed.connect(table_view._configure_headers)
            table_view._configure_headers()
            
            # Clear UI state
            self._details_bar.clear()
            self._card_designer.clear()
            self._card_table_widget._remove_row_btn.setEnabled(False)
            
            # Reconnect model signals for unsaved changes tracking
            self._model.dataChanged.connect(self._mark_unsaved)
            self._model.template_changed.connect(self._mark_unsaved)
            
            self._update_title()
            self.statusBar().showMessage(f"Project '{name}' opened.", 3000)
            return True
        except (OSError, json.JSONDecodeError, KeyError) as e:
            QMessageBox.critical(self, "Error", f"Failed to load: {e}")
            return False
    
    def _list_projects(self) -> list[str]:
        if not self.BASE_DIR.exists():
            return []
        
        projects = []
        for item in self.BASE_DIR.iterdir():
            if item.is_dir() and (item / "project.json").exists():
                projects.append(item.name)
        return sorted(projects)
    
    def _on_export(self) -> None:
        if self._model.rowCount() == 0:
            QMessageBox.information(self, "No Cards", "Add some cards before exporting.")
            return
        
        dialog = ExportOptionsDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            export_format = dialog.get_format()
            cards_per_page = dialog.get_cards_per_page()
            
            if export_format == "PDF":
                file_filter = "PDF Files (*.pdf)"
                default_ext = ".pdf"
            else:
                file_filter = "Word Documents (*.docx)"
                default_ext = ".docx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Business Cards",
                f"business_cards{default_ext}", file_filter
            )
            
            if not file_path:
                return
            
            output_path = Path(file_path)
            if not output_path.suffix.lower() == default_ext:
                output_path = output_path.with_suffix(default_ext)
            
            try:
                from src.business_card_generator.core.export_engine import ExportEngine
                
                project_path = None
                if self._current_project_name:
                    project_path = str(self.BASE_DIR / self._current_project_name)
                
                engine = ExportEngine(self._model, project_path)
                
                if export_format == "PDF":
                    success = engine.export_pdf(output_path, cards_per_page)
                else:
                    success = engine.export_docx(output_path, cards_per_page)
                
                if success:
                    reply = QMessageBox.information(
                        self, "Export Complete",
                        f"Exported to:\n{output_path}\n\nOpen file location?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    if reply == QMessageBox.StandardButton.Yes:
                        QDesktopServices.openUrl(QUrl.fromLocalFile(str(output_path.parent)))
                else:
                    QMessageBox.warning(self, "Export Failed", "No cards to export.")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Export failed: {e}")
    
    def _on_import_excel(self) -> None:
        """Import card data from an Excel file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import from Excel",
            "", "Excel Files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Column{col}")
            
            if not headers:
                QMessageBox.warning(self, "Import Error", "No columns found in Excel file.")
                return
            
            # Image file extensions for auto-detection
            image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}
            
            # Detect which columns are likely image columns based on content
            image_columns = set()
            for col_idx, header in enumerate(headers):
                # Check first few data rows to detect image columns
                for row in range(2, min(ws.max_row + 1, 10)):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value:
                        value_str = str(cell_value).lower()
                        # Check if value looks like an image filename
                        if any(value_str.endswith(ext) for ext in image_extensions):
                            image_columns.add(col_idx)
                            break
                        # Also check if header suggests image
                        header_lower = header.lower()
                        if any(kw in header_lower for kw in ['photo', 'image', 'logo', 'picture', 'pic']):
                            image_columns.add(col_idx)
                            break
            
            # Clear existing template fields and create new ones from Excel columns
            self._model.beginResetModel()
            self._model._template.fields.clear()
            self._model._rows.clear()
            
            from src.business_card_generator.models.card import FieldDefinition, FieldType, CardRow
            
            # Create fields from headers
            for col_idx, header in enumerate(headers):
                field_type = FieldType.IMAGE if col_idx in image_columns else FieldType.TEXT
                
                # Calculate default position
                y_offset = col_idx * 25
                field_def = FieldDefinition(
                    name=header,
                    field_type=field_type,
                    x=10, y=10 + y_offset,
                    width=150 if field_type == FieldType.TEXT else 80,
                    height=25 if field_type == FieldType.TEXT else 80
                )
                self._model._template.fields.append(field_def)
            
            # Import data rows
            for row_num in range(2, ws.max_row + 1):
                row_data = CardRow()
                has_data = False
                
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row_num, column=col_idx + 1).value
                    if cell_value is not None:
                        has_data = True
                        field_id = self._model._template.fields[col_idx].id
                        row_data.set_value(field_id, str(cell_value))
                
                if has_data:
                    self._model._rows.append(row_data)
            
            self._model.endResetModel()
            self._model.template_changed.emit()
            
            # Update UI
            self._card_designer._rebuild_field_widgets()
            self._card_designer.clear()
            self._details_bar.clear()
            self._mark_unsaved()
            
            # Show summary
            num_rows = len(self._model._rows)
            num_cols = len(headers)
            image_col_names = [headers[i] for i in image_columns]
            
            msg = f"Imported {num_rows} cards with {num_cols} fields."
            if image_col_names:
                msg += f"\n\nDetected image columns: {', '.join(image_col_names)}"
            
            QMessageBox.information(self, "Import Complete", msg)
            self.statusBar().showMessage(f"Imported {num_rows} cards from Excel.", 3000)
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import: {e}")
    
    def closeEvent(self, event: QCloseEvent) -> None:
        if self._has_unsaved_changes:
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                "You have unsaved changes. Exit anyway?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                event.ignore()
                return
        event.accept()
    
    # Accessors for testing
    def get_model(self) -> CardDataModel:
        return self._model
    
    def get_card_designer(self) -> CardDesigner:
        return self._card_designer
    
    def get_card_table_view(self) -> CardTableWidget:
        return self._card_table_widget
    
    def get_details_bar(self) -> DetailsBar:
        return self._details_bar
    
    def get_splitter(self) -> QSplitter:
        return self._splitter
    
    def has_unsaved_changes(self) -> bool:
        return self._has_unsaved_changes
    
    def set_unsaved_changes(self, value: bool) -> None:
        self._has_unsaved_changes = value
        self._update_title()
    
    def get_menu_action(self, name: str) -> Optional[QAction]:
        return self._menu_actions.get(name)


class OpenProjectDialog(QDialog):
    """Dialog for selecting a project to open."""
    
    def __init__(self, projects: list[str], parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("Open Project")
        self.setMinimumSize(300, 200)
        
        self._selected: Optional[str] = None
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Select a project:"))
        
        self._list = QListWidget()
        self._list.addItems(projects)
        self._list.itemDoubleClicked.connect(self._on_double_click)
        layout.addWidget(self._list)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _on_double_click(self, item) -> None:
        self._selected = item.text()
        self.accept()
    
    def _on_accept(self) -> None:
        item = self._list.currentItem()
        if item:
            self._selected = item.text()
            self.accept()
    
    def get_selected_project(self) -> Optional[str]:
        return self._selected


class ExportOptionsDialog(QDialog):
    """Dialog for selecting export options."""
    
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("Export Options")
        self.setMinimumWidth(300)
        
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self._format_combo = QComboBox()
        self._format_combo.addItems(["PDF", "Word (DOCX)"])
        form.addRow("Format:", self._format_combo)
        
        self._cards_per_page = QComboBox()
        self._cards_per_page.addItems(["1", "2", "4", "6", "8", "10"])
        self._cards_per_page.setCurrentText("4")
        self._cards_per_page.setToolTip("Number of business cards to fit on each page")
        form.addRow("Cards per page:", self._cards_per_page)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def get_format(self) -> str:
        text = self._format_combo.currentText()
        return "PDF" if "PDF" in text else "DOCX"
    
    def get_cards_per_page(self) -> int:
        return int(self._cards_per_page.currentText())
